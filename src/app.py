import os
import io
import re
import boto3
from openpyxl import load_workbook, Workbook 

s3_client = boto3.client('s3')
dynamodb = boto3.resource('dynamodb')

DESTINATION_BUCKET = os.environ.get('DESTINATION_BUCKET')
TABLE_NAME = os.environ.get('TABLE_NAME')
table = dynamodb.Table(TABLE_NAME)

# Lambda handler function for processing the uploaded Excel file handled by S3 event . An Excel file of Translation is read, cleaned by looking for '(not used)' patterns, and stored in DynamoDB and a different S3. 
def lambda_handler(event, context):
    try:
        source_bucket = event['Records'][0]['s3']['bucket']['name']
        key = event['Records'][0]['s3']['object']['key']
        

        filename = key.split('/')[-1]
        project_code = filename.split('_')[0] 
        

        response = s3_client.get_object(Bucket=source_bucket, Key=key)
        file_stream = io.BytesIO(response['Body'].read()) 
        
        wb = load_workbook(filename=file_stream, data_only=True)
        ws = wb.active 
        not_used_pattern = re.compile(r'\(not used\)', re.IGNORECASE)
        

        clean_wb = Workbook()
        clean_ws = clean_wb.active
        clean_ws.append(['Language ID', 'Form ID', 'Form name', 'Comments', 'text'])
        
        processed_count = 0
        

        with table.batch_writer() as batch:
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                lang_id, form_id, form_name, comments, text = row
                
                form_name_str = str(form_name) if form_name is not None else ""
                comments_str = str(comments) if comments is not None else ""
                text_str = str(text) if text is not None else ""
                

                if not_used_pattern.search(form_name_str) or not_used_pattern.search(comments_str):
                    continue
                    
                clean_ws.append([lang_id, form_id, form_name_str, comments_str, text_str])
                

                batch.put_item(Item={
                    'PK': f"PROJECT#{project_code}",
                    'SK': f"{lang_id}#{form_id}#{str(idx).zfill(4)}",
                    'Language_id': lang_id,
                    'Form_Name': form_name_str,
                    'Comments': comments_str,
                    'Text': text_str
                })
                processed_count += 1
                

        output_stream = io.BytesIO()
        clean_wb.save(output_stream)
        
        clean_key = f"cleaned_{filename}"
        s3_client.put_object(
            Bucket=DESTINATION_BUCKET,
            Key=clean_key,
            Body=output_stream.getvalue()
        )
            
        print(f"Successfully processed {processed_count} records for {project_code}")
        return {'statusCode': 200, 'body': f"Success: {processed_count} records processed for {project_code}."}

    except Exception as e:
        print(f"Error processing {key}: {str(e)}")
        raise e